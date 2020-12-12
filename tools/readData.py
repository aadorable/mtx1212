# *** coding: utf-8 ***
#@Time   : 2020/11/26 17:48
#@Author : xueqing.wu
#@Email  : wuxueqing@126.com
#@File   : readData.py

import openpyxl
import yaml
import csv
from settings import DIR_NAME

class ReadData():
    def get_excel_list(self, filename):
        wb = openpyxl.load_workbook(DIR_NAME + '/data/%s' % filename)
        ws = wb['测试用例']
        all_cases = []
        select_data_area = ws.iter_rows(min_col=1, max_col=ws.max_column, min_row=2, max_row=ws.max_row)
        for rows in select_data_area:
            case_list = [cell.value for cell in rows]
            all_cases.append(case_list)
        return all_cases

    def get_excel_dict(self, header, filename):
        wb = openpyxl.load_workbook(DIR_NAME + '/data/%s' % filename)
        ws = wb['测试用例']
        all_cases = []
        for row in range(2, ws.max_row+1):
            case_dict = {}
            col = 1
            for key in header:
                case_dict[key] = ws.cell(row, col).value
                col += 1
            all_cases.append(case_dict)
        return all_cases

    def get_yaml(self,filename,key):
        with open(DIR_NAME + '/data/%s' % filename, encoding='utf-8') as f:
            # 最新的知识点
            yaml_data = yaml.safe_load(f)
            # print(yaml_data)
            cases_dict = yaml_data.get(key)
            # case_object是可迭代对象  for ...in..,list(),列表.extend()
            case_object = cases_dict.values()
            # print(case_object)
            # data_list.extend(case_object)
            data = list(case_object)
            return data

    def get_csv(self, filename):
        with open(DIR_NAME + '/data/%s' % filename, encoding='utf-8') as f:
            next(f)
            csv_data = csv.reader(f)
            return list(csv_data)



if __name__ == '__main__':
    print(ReadData().get_excel_list('login_data.xlsx'))
    print(ReadData().get_excel_list('reg_data.xlsx'))
    print(ReadData().get_excel_dict(['username', 'password', 'exp'], 'login_data.xlsx'))